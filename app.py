from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import List
import openpyxl

class Race(BaseModel):
    year: int
    winner: str
    country: str

app = FastAPI()
wb = openpyxl.load_workbook("data/tour_de_france.xlsx")
sheet = wb.active

def row_to_race(row):
    return Race(year=row[0].value, winner=row[1].value, country=row[2].value)

# Endpoint to retrieve all races
@app.get("/Race/", response_model=List[Race])
async def read_races():
    races = [row_to_race(row) for row in sheet.iter_rows(min_row=2, values_only=True)]
    return races

# Endpoint to retrieve a specific race by year
@app.get("/Race/{year}", response_model=Race)
async def read_race(year: int):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == year:
            return row_to_race(row)

# Endpoint to create a new race
@app.post("/Race/")
async def create_race(race: Race):
    sheet.append([race.year, Race.winner, Race.country])
    wb.save("data/tour_de_france.xlsx")
    return Race

# Endpoint to update an existing race
@app.put("/Race/{year}")
async def update_race(year: int, Race: Race):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == year:
            row[1] = Race.winner
            row[2] = Race.country
            wb.save("data/tour_de_france.xlsx")
            return Race
    raise HTTPException(status_code=404, detail="Race not found")

# Endpoint to delete a race
@app.delete("/Race/{year}")
async def delete_race(year: int):
    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if row[0] == year:
            sheet.delete_race(idx)
            wb.save("data/tour_de_france.xlsx")
            return {"message": "Race deleted successfully"}
    raise HTTPException(status_code=404, detail="Race not found")
